"""
plan-takeoff backend — PDF upload, conversion, serving, and export.

Coordinate System & Resolution Notes
=====================================
- PDFs are rasterised at **300 DPI** using PyMuPDF (fitz).
- The conversion matrix is `fitz.Matrix(300/72, 300/72)` which scales the
  default 72-DPI PDF user-space by ~4.1667×.
- A PDF point at (x_pt, y_pt) maps to pixel (x_px, y_px) via:
      x_px = x_pt * (300 / 72)
      y_px = y_pt * (300 / 72)
- PNGs are saved as RGBA with no additional scaling or compression artefacts,
  so every pixel in the served image corresponds exactly to the above mapping.
- The frontend canvas must use the **natural image dimensions** (PNG width ×
  height) as its coordinate space so that overlay coordinates can be translated
  back to PDF points using the inverse ratio (72 / 300).
"""

from __future__ import annotations

import io
import json
import math
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from PIL import Image, ImageDraw
from pydantic import BaseModel

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DPI = 300
"""Rasterisation resolution. Change this constant and re-upload to adjust."""

UPLOAD_DIR = Path(__file__).resolve().parent.parent / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

app = FastAPI(title="plan-takeoff")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # tighten in production
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# In-memory storage for room measurements per job
# ---------------------------------------------------------------------------

# {job_id: {"rooms": [...], "scale": float, "ceiling_height": float, "filename": str}}
_job_data: dict[str, dict[str, Any]] = {}

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------


class WallSegment(BaseModel):
    start_x: float  # pixel coords
    start_y: float
    end_x: float
    end_y: float
    length_mm: float


class Room(BaseModel):
    label: str
    page: int
    polygon: list[list[float]]  # [[x, y], ...] in pixel coords
    wall_segments: list[WallSegment]
    floor_area_m2: float
    total_wall_length_mm: float
    total_wall_area_m2: float
    perimeter_mm: float


class RoomData(BaseModel):
    rooms: list[Room]
    scale: float = 1.0  # pixels per mm
    ceiling_height_mm: float = 2400.0
    filename: str = ""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

ZOOM = DPI / 72  # 72 DPI is the PDF default user-space unit


def _convert_pdf(pdf_path: Path, out_dir: Path) -> int:
    """Render every page of *pdf_path* to a PNG in *out_dir*.

    Returns the total number of pages.
    """
    doc = fitz.open(str(pdf_path))
    mat = fitz.Matrix(ZOOM, ZOOM)
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        pix.save(str(out_dir / f"page_{page_num}.png"))
    page_count = len(doc)
    doc.close()
    return page_count


def _validate_job(job_id: str) -> Path:
    """Return job directory or raise 404."""
    job_dir = UPLOAD_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="Job not found.")
    return job_dir


def _get_rooms(job_id: str) -> list[dict[str, Any]]:
    """Return rooms list for a job (may be empty)."""
    data = _job_data.get(job_id, {})
    return data.get("rooms", [])


def _composite_image(job_dir: Path, page: int, rooms: list[dict[str, Any]]) -> Image.Image:
    """Load a page PNG and draw room polygon overlays onto it."""
    img_path = job_dir / f"page_{page}.png"
    if not img_path.exists():
        raise HTTPException(status_code=404, detail=f"Page {page} image not found.")
    img = Image.open(img_path).convert("RGBA")

    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    colors = [
        (79, 70, 229, 50),   # indigo
        (16, 185, 129, 50),  # emerald
        (245, 158, 11, 50),  # amber
        (239, 68, 68, 50),   # red
        (139, 92, 246, 50),  # violet
        (6, 182, 212, 50),   # cyan
    ]
    border_colors = [
        (79, 70, 229, 200),
        (16, 185, 129, 200),
        (245, 158, 11, 200),
        (239, 68, 68, 200),
        (139, 92, 246, 200),
        (6, 182, 212, 200),
    ]

    page_rooms = [r for r in rooms if r.get("page") == page]
    for i, room in enumerate(page_rooms):
        polygon = room.get("polygon", [])
        if len(polygon) < 3:
            continue
        pts = [(p[0], p[1]) for p in polygon]
        color_idx = i % len(colors)
        draw.polygon(pts, fill=colors[color_idx], outline=border_colors[color_idx])

        # Draw label at centroid
        cx = sum(p[0] for p in pts) / len(pts)
        cy = sum(p[1] for p in pts) / len(pts)
        label = room.get("label", f"Room {i + 1}")
        draw.text((cx, cy), label, fill=(0, 0, 0, 220))

    composited = Image.alpha_composite(img, overlay)
    return composited.convert("RGB")


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@app.post("/api/upload")
async def upload_pdf(file: UploadFile):
    """Accept a PDF upload, rasterise pages at 300 DPI, return job metadata."""
    if file.content_type not in ("application/pdf",):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    job_id = uuid.uuid4().hex
    job_dir = UPLOAD_DIR / job_id
    job_dir.mkdir(parents=True)

    # Persist the original PDF
    pdf_path = job_dir / "original.pdf"
    contents = await file.read()
    pdf_path.write_bytes(contents)

    # Convert to PNGs
    page_count = _convert_pdf(pdf_path, job_dir)

    # Initialise job data with filename
    _job_data[job_id] = {
        "rooms": [],
        "scale": 1.0,
        "ceiling_height_mm": 2400.0,
        "filename": file.filename or "unknown.pdf",
    }

    return {
        "job_id": job_id,
        "page_count": page_count,
        "dpi": DPI,
        "pages": [
            {"page": i, "url": f"/api/images/{job_id}/{i}"}
            for i in range(page_count)
        ],
    }


@app.get("/api/images/{job_id}/{page}")
async def get_page_image(job_id: str, page: int):
    """Serve a rasterised page PNG."""
    img_path = UPLOAD_DIR / job_id / f"page_{page}.png"
    if not img_path.exists():
        raise HTTPException(status_code=404, detail="Image not found.")
    return FileResponse(img_path, media_type="image/png")


@app.get("/api/pdf/{job_id}")
async def get_pdf(job_id: str):
    """Serve the original uploaded PDF."""
    pdf_path = UPLOAD_DIR / job_id / "original.pdf"
    if not pdf_path.exists():
        raise HTTPException(status_code=404, detail="PDF not found.")
    return FileResponse(pdf_path, media_type="application/pdf")


# ---------------------------------------------------------------------------
# Room data endpoints
# ---------------------------------------------------------------------------


@app.post("/api/jobs/{job_id}/rooms")
async def save_rooms(job_id: str, data: RoomData):
    """Save room measurement data for a job."""
    _validate_job(job_id)
    existing = _job_data.get(job_id, {})
    _job_data[job_id] = {
        "rooms": [r.model_dump() for r in data.rooms],
        "scale": data.scale,
        "ceiling_height_mm": data.ceiling_height_mm,
        "filename": data.filename or existing.get("filename", ""),
    }
    return {"status": "ok", "room_count": len(data.rooms)}


@app.get("/api/jobs/{job_id}/rooms")
async def get_rooms(job_id: str):
    """Retrieve room measurement data for a job."""
    _validate_job(job_id)
    data = _job_data.get(job_id, {})
    return {
        "rooms": data.get("rooms", []),
        "scale": data.get("scale", 1.0),
        "ceiling_height_mm": data.get("ceiling_height_mm", 2400.0),
        "filename": data.get("filename", ""),
    }


# ---------------------------------------------------------------------------
# Export endpoints
# ---------------------------------------------------------------------------


@app.get("/api/export/csv/{job_id}")
async def export_csv(job_id: str):
    """Export room measurements as an Excel (.xlsx) spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    _validate_job(job_id)
    data = _job_data.get(job_id, {})
    rooms = data.get("rooms", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Room Measurements"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    # Summary sheet headers
    headers = [
        "Room Label",
        "Floor Area (m²)",
        "Total Wall Length (mm)",
        "Total Wall Area (m²)",
        "Perimeter (mm)",
        "Page",
        "Wall Count",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row_idx, room in enumerate(rooms, 2):
        ws.cell(row=row_idx, column=1, value=room.get("label", ""))
        ws.cell(row=row_idx, column=2, value=round(room.get("floor_area_m2", 0), 3))
        ws.cell(row=row_idx, column=3, value=round(room.get("total_wall_length_mm", 0), 1))
        ws.cell(row=row_idx, column=4, value=round(room.get("total_wall_area_m2", 0), 3))
        ws.cell(row=row_idx, column=5, value=round(room.get("perimeter_mm", 0), 1))
        ws.cell(row=row_idx, column=6, value=room.get("page", 0) + 1)
        ws.cell(row=row_idx, column=7, value=len(room.get("wall_segments", [])))

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    # Wall segments detail sheet
    ws2 = wb.create_sheet("Wall Segments")
    seg_headers = [
        "Room Label",
        "Wall #",
        "Start X (px)",
        "Start Y (px)",
        "End X (px)",
        "End Y (px)",
        "Length (mm)",
    ]
    for col, h in enumerate(seg_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    seg_row = 2
    for room in rooms:
        for wall_idx, seg in enumerate(room.get("wall_segments", []), 1):
            ws2.cell(row=seg_row, column=1, value=room.get("label", ""))
            ws2.cell(row=seg_row, column=2, value=wall_idx)
            ws2.cell(row=seg_row, column=3, value=round(seg.get("start_x", 0), 1))
            ws2.cell(row=seg_row, column=4, value=round(seg.get("start_y", 0), 1))
            ws2.cell(row=seg_row, column=5, value=round(seg.get("end_x", 0), 1))
            ws2.cell(row=seg_row, column=6, value=round(seg.get("end_y", 0), 1))
            ws2.cell(row=seg_row, column=7, value=round(seg.get("length_mm", 0), 1))
            seg_row += 1

    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = data.get("filename", "export").replace(".pdf", "")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}_takeoff.xlsx"'
        },
    )


@app.get("/api/export/json/{job_id}")
async def export_json(job_id: str):
    """Export all raw measurement data as JSON."""
    _validate_job(job_id)
    data = _job_data.get(job_id, {})
    export = {
        "job_id": job_id,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "filename": data.get("filename", ""),
        "scale_px_per_mm": data.get("scale", 1.0),
        "ceiling_height_mm": data.get("ceiling_height_mm", 2400.0),
        "dpi": DPI,
        "rooms": data.get("rooms", []),
    }
    content = json.dumps(export, indent=2)
    filename = data.get("filename", "export").replace(".pdf", "")
    return Response(
        content=content,
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}_takeoff.json"'
        },
    )


@app.get("/api/export/pdf/{job_id}")
async def export_pdf(job_id: str):
    """Generate a PDF report with composited plan images and measurement tables."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image as RLImage,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    job_dir = _validate_job(job_id)
    data = _job_data.get(job_id, {})
    rooms = data.get("rooms", [])
    filename = data.get("filename", "unknown.pdf")
    scale = data.get("scale", 1.0)
    ceiling_height_mm = data.get("ceiling_height_mm", 2400.0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=18,
        spaceAfter=6 * mm,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=8 * mm,
        spaceAfter=4 * mm,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=2 * mm,
    )

    story: list[Any] = []

    # Title
    story.append(Paragraph("Plan Take-Off Report", title_style))

    # Project metadata
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    meta_lines = [
        f"<b>File:</b> {filename}",
        f"<b>Date:</b> {now}",
        f"<b>Scale:</b> {scale:.4f} px/mm",
        f"<b>Ceiling Height:</b> {ceiling_height_mm:.0f} mm",
        f"<b>DPI:</b> {DPI}",
        f"<b>Job ID:</b> {job_id[:16]}…",
    ]
    for line in meta_lines:
        story.append(Paragraph(line, meta_style))
    story.append(Spacer(1, 6 * mm))

    # Collect unique pages that have rooms
    pages_with_rooms = sorted(set(r.get("page", 0) for r in rooms))
    if not pages_with_rooms:
        # If no rooms, show page 0
        pages_with_rooms = [0]

    # Composited plan images
    usable_width = A4[0] - 30 * mm
    for page_num in pages_with_rooms:
        story.append(Paragraph(f"Page {page_num + 1}", heading_style))

        try:
            comp = _composite_image(job_dir, page_num, rooms)
            img_buf = io.BytesIO()
            comp.save(img_buf, format="PNG")
            img_buf.seek(0)

            # Scale image to fit page width
            orig_w, orig_h = comp.size
            aspect = orig_h / orig_w
            img_w = usable_width
            img_h = img_w * aspect

            # Cap height to avoid overflow
            max_h = A4[1] - 80 * mm
            if img_h > max_h:
                img_h = max_h
                img_w = img_h / aspect

            rl_img = RLImage(img_buf, width=img_w, height=img_h)
            story.append(rl_img)
        except Exception:
            story.append(Paragraph(f"<i>Could not render page {page_num + 1}</i>", styles["Normal"]))

        story.append(Spacer(1, 6 * mm))

    # Measurement summary table
    if rooms:
        story.append(Paragraph("Room Measurements", heading_style))

        table_data = [
            [
                "Room",
                "Floor Area\n(m²)",
                "Wall Length\n(mm)",
                "Wall Area\n(m²)",
                "Perimeter\n(mm)",
                "Page",
            ]
        ]
        for room in rooms:
            table_data.append([
                room.get("label", ""),
                f"{room.get('floor_area_m2', 0):.2f}",
                f"{room.get('total_wall_length_mm', 0):.0f}",
                f"{room.get('total_wall_area_m2', 0):.2f}",
                f"{room.get('perimeter_mm', 0):.0f}",
                str(room.get("page", 0) + 1),
            ])

        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)

        # Wall segment details
        story.append(Paragraph("Wall Segment Details", heading_style))
        for room in rooms:
            segments = room.get("wall_segments", [])
            if not segments:
                continue
            story.append(Paragraph(f"<b>{room.get('label', 'Room')}</b>", styles["Normal"]))
            seg_data = [["Wall #", "Length (mm)"]]
            for idx, seg in enumerate(segments, 1):
                seg_data.append([str(idx), f"{seg.get('length_mm', 0):.1f}"])
            seg_tbl = Table(seg_data, colWidths=[60, 80])
            seg_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(seg_tbl)
            story.append(Spacer(1, 3 * mm))

    else:
        story.append(Paragraph("No room measurements recorded.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)

    export_filename = filename.replace(".pdf", "")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{export_filename}_report.pdf"'
        },
    )
